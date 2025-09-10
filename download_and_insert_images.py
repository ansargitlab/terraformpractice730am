"""
download_and_insert_images.py

Run this locally (on your PC) to download the first image from each Bing Images search URL placed in column K
and insert that image into the same cell (column K) for each row in both sheets.

Requirements (install via pip):
    pip install openpyxl requests pillow beautifulsoup4 lxml

Usage:
    python download_and_insert_images.py "Iccubators & centrifuge summary_with_image_links.xlsx" "output_with_images.xlsx"

Notes:
- This script fetches the Bing Images search result page for each URL, picks the first image result, downloads it,
  and embeds it into the Excel file using openpyxl. Do not run this on a restricted network.
- If an image cannot be downloaded, the URL remains in the cell and a message is printed.
"""

import sys
import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from bs4 import BeautifulSoup
from PIL import Image as PILImage

def download_first_image_from_bing(search_url, headers):
    try:
        resp = requests.get(search_url, headers=headers, timeout=15)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, "lxml")
        # Bing image results include <a class="iusc" ... m="{... 'murl':'<url>' ...}">
        # We'll try to find images via 'm' attribute or fallback to <img> tags
        a_iusc = soup.find("a", {"class":"iusc"})
        if a_iusc and a_iusc.get("m"):
            import json
            try:
                mdata = json.loads(a_iusc["m"])
                if "murl" in mdata:
                    return mdata["murl"]
            except Exception:
                pass
        # fallback: first <img> inside results
        img = soup.find("img")
        if img and img.get("src"):
            return img["src"]
        return None
    except Exception as e:
        print("Download error:", e)
        return None

def embed_image_in_cell(ws, row, col, image_bytes):
    # Save to a BytesIO and create openpyxl Image
    bio = BytesIO(image_bytes)
    try:
        pil = PILImage.open(bio)
        # Optionally resize if too large
        max_width, max_height = 600, 400
        w, h = pil.size
        scale = min(1, max_width / w, max_height / h)
        if scale < 1:
            pil = pil.resize((int(w*scale), int(h*scale)), PILImage.LANCZOS)
        out = BytesIO()
        pil.save(out, format="PNG")
        out.seek(0)
    except Exception:
        # if Pillow fails, use raw bytes
        out = BytesIO(image_bytes)
        out.seek(0)
    img = XLImage(out)
    cell = ws.cell(row=row, column=col)
    # place image anchored to the cell
    img.anchor = f"{cell.column_letter}{row}"
    ws.add_image(img)
    # adjust row height (approx)
    try:
        ws.row_dimensions[row].height = 80
    except Exception:
        pass

def main():
    if len(sys.argv) < 3:
        print("Usage: python download_and_insert_images.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    input_xlsx = sys.argv[1]
    output_xlsx = sys.argv[2]
    wb = load_workbook(input_xlsx)
    headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in range(2, ws.max_row+1):
            url_cell = ws.cell(row=row, column=11)  # Column K
            url = url_cell.value
            if not url or not str(url).startswith("http"):
                continue
            print(f"Processing {sheetname} row {row} -> {url}")
            image_url = download_first_image_from_bing(url, headers)
            if not image_url:
                print("  No image URL found for search page.")
                continue
            try:
                r = requests.get(image_url, headers=headers, timeout=15)
                if r.status_code == 200:
                    embed_image_in_cell(ws, row, 11, r.content)
                    # Optionally clear the URL text after embedding; comment out if you want to keep it
                    # ws.cell(row=row, column=11).value = None
                else:
                    print("  Failed to download image:", r.status_code, image_url)
            except Exception as e:
                print("  Exception downloading image:", e)
    wb.save(output_xlsx)
    print("Done. Saved to", output_xlsx)

if __name__ == "__main__":
    main()
